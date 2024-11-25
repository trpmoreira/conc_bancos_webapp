import os
import sys
import shutil
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
from base_dados import get_bank_transactions

# Lista com os nomes dos meses em português
meses_pt = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]

# Mapeamento entre as abas e os ficheiros correspondentes
mapeamento = {
    "STD DO": "120101",
    "STD CR": "120102",
    "BCP DO": "120301",
    "BCP 2": "120302",
    "BCP 3": "120303",
    "BCP 4": "120304",
    "Montepio": "120401",
    "BIC": "120501",
    "CGD": "120601"
}

# Adicionar no início do arquivo junto com os outros mapeamentos
mapeamento_colunas_valor = {
    "STD DO": "Montante",  # Coluna E
    "STD CR": "Montante",
    "BCP DO": "Valor",     # Coluna H
    "BCP 2": "Valor",
    "BCP 3": "Valor",
    "BCP 4": "Valor",
    "Montepio": "IMPORTÂNCIA",   # Ajustar nome correto
    "BIC": "Valor",        # Ajustar nome correto
    "CGD": "Montante "         # Ajustar nome correto
}

def validate_document_format(doc, ano, mes, banco_code):
    if not isinstance(doc, str):
        return False

    try:
        # Verifica se tem o tamanho correto (10 caracteres)


        # Verifica se os primeiros dois dígitos são o ano
        ano_doc = doc[1:3]
        if ano_doc != str(ano)[-2:]:  # Pega os últimos 2 dígitos do ano
            return False

        # Verifica se os próximos dois dígitos são o mês
        mes_doc = doc[3:5]
        if mes_doc != f"{mes:02d}":  # Formata o mês com 2 dígitos
            return False

        # Verifica se os próximos dois dígitos são o código do banco
        banco_doc = doc[5:7]
        if banco_doc != banco_code:
            return False

        # Verifica se os últimos 4 dígitos são números
        num_doc = doc[7:11]
        if not num_doc.isdigit():
            return False

        return True
    except:
        return False

def create_invalid_docs_report(dfs_phc, ano, mes):
    invalid_docs = []

    banco_codes = {
        "120101": "01",
        "120102": "02",
        "120103": "03",
        "120301": "04",
        "120302": "05",
        "120401": "06",
        "120201": "07",
        "120601": "08",
        "120501": "09",
        "120303": "12",
        "120304": "13"
    }

    for conta, df in dfs_phc.items():
        banco_code = banco_codes.get(conta)
        if not banco_code:
            continue

        for _, row in df.iterrows():
            doc = row.get('Documento')
            if doc:
                doc_limpo = clean_cell_value(str(doc))
                if not validate_document_format(doc_limpo, ano, mes, banco_code):
                    invalid_docs.append({
                        'Conta': clean_cell_value(conta),
                        'Nº': clean_cell_value(row.get('Nº', '')),
                        'Documento': doc_limpo
                    })

    df_invalid = pd.DataFrame(invalid_docs)
    return df_invalid

def clean_cell_value(value):
    if value is None:
        return ""

    # Converte para string e remove espaços extras
    value = str(value).strip()

    # Se parece ser um número de documento (começa com B e tem números)
    if value.startswith('B') and any(c.isdigit() for c in value):
        # Remove todos os espaços e pega apenas os primeiros 11 caracteres
        value = ''.join(value.split())[:11]

    # Lista de caracteres ilegais no Excel
    illegal_chars = [
        '\x00', '\x01', '\x02', '\x03', '\x04', '\x05', '\x06', '\x07', '\x08', '\x0b',
        '\x0c', '\x0e', '\x0f', '\x10', '\x11', '\x12', '\x13', '\x14', '\x15', '\x16',
        '\x17', '\x18', '\x19', '\x1a', '\x1b', '\x1c', '\x1d', '\x1e', '\x1f'
    ]

    # Remove caracteres ilegais
    for char in illegal_chars:
        value = value.replace(char, '')

    # Remove outros caracteres problemáticos
    problematic_chars = [':', '\\', '/', '?', '*', '[', ']', '\t', '\n', '\r']
    for char in problematic_chars:
        value = value.replace(char, '')

    # Se ainda houver caracteres não imprimíveis, substitui por espaço
    cleaned_value = ''.join(char if char.isprintable() else ' ' for char in value)

    return cleaned_value.strip()

def create_and_update_files(month_int):
    # Verifica se o número do mês é válido
    if month_int < 1 or month_int > 12:
        print("Por favor, insira um número de mês válido (1-12).")
        return

    # Obtém o nome do mês a partir do número
    month_name = meses_pt[month_int]

    # Nome da pasta: "int - mês"
    folder_name = f"{month_int} - {month_name}"

    # Caminho para o diretório atual e para os ficheiros
    base_dir = Path.cwd()
    bancos_dir = base_dir / "Bancos"
    template_file = base_dir / "Layout Bancos.xlsx"
    bancos_file = next(bancos_dir.glob(f"{month_int:02d} - Bancos*.xlsx"), None)

    # Verifica se os ficheiros necessários existem
    if not template_file.exists():
        print("O ficheiro modelo 'Layout Bancos.xlsx' não foi encontrado.")
        return

    if not bancos_file.exists():
        print("O ficheiro 'bancos.xlsx' não foi encontrado.")
        return

    if not bancos_dir.exists():
        print("A pasta 'Bancos' não foi encontrada. Certifique-se de que existe na diretoria atual.")
        return

    if not bancos_file:
        print(f"O ficheiro para o mês {month_name} ('{month_int:02d} - Bancos') não foi encontrado na pasta 'Bancos'.")
        return

    # Cria a pasta se não existir
    folder_path = base_dir / folder_name
    folder_path.mkdir(exist_ok=True)

    # Nomes dos ficheiros e estrutura de nomenclatura
    file_names = [
        "120101 - Santander DO",
        "120102 - Santander CR",
        "120301 - BCP DO",
        "120302 - BCP 2",
        "120303 - BCP 3",
        "120304 - BCP 4",
        "120401 - Montepio DO",
        "120501 - BIC DO",
        "120601 - CGD DO"
    ]

    # Criação de cada ficheiro com cópia do modelo
    created_files = []
    for name in file_names:
        file_path = folder_path / f"{name} - {month_name}.xlsx"
        shutil.copy(template_file, file_path)
        print(f"Ficheiro criado: {file_path}")
        created_files.append(file_path)

    try:
        # Carrega os workbooks
        wb_bancos = load_workbook(bancos_file)

        # Busca dados do PHC para todas as contas
        dfs_phc = {}
        for prefixo in mapeamento.values():
            df = get_bank_transactions(prefixo, month_int)
            if df is not None:
                dfs_phc[prefixo] = df

        # Processar dados dos bancos
        for aba, prefixo in mapeamento.items():
            if aba in wb_bancos.sheetnames:
                # Dados da aba do banco
                sheet_banco = wb_bancos[aba]
                dados_banco = [[cell.value for cell in row] for row in sheet_banco.iter_rows()]

                # Localiza o ficheiro correspondente para bancos
                file_path_banco = next(
                    (f for f in created_files if f.name.startswith(f"{prefixo}")),
                    None
                )

                if file_path_banco:
                    wb_destino_banco = load_workbook(file_path_banco)

                    # Certifique-se de que a aba 'Banco' existe no ficheiro destino
                    if "Banco" in wb_destino_banco.sheetnames:
                        sheet_destino_banco = wb_destino_banco["Banco"]

                        # Copia os dados dos bancos
                        for i, linha in enumerate(dados_banco, start=1):
                            for j, valor in enumerate(linha, start=1):
                                valor_limpo = clean_cell_value(valor)
                                sheet_destino_banco.cell(row=i, column=j, value=valor_limpo)

                        wb_destino_banco.save(file_path_banco)
                        print(f"Dados bancários atualizados em: {file_path_banco}")

        # Processar dados do PHC (nova implementação)
        for prefixo, df_phc in dfs_phc.items():
            file_path_phc = next(
                (f for f in created_files if f.name.startswith(f"{prefixo}")),
                None
            )

            if file_path_phc:
                wb_destino_phc = load_workbook(file_path_phc)

                if "PHC" in wb_destino_phc.sheetnames:
                    sheet_destino_phc = wb_destino_phc["PHC"]

                    # Escreve o cabeçalho
                    for j, coluna in enumerate(df_phc.columns, start=1):
                        sheet_destino_phc.cell(row=1, column=j, value=coluna)

                    # Escreve os dados
                    for i, row in enumerate(df_phc.values, start=2):
                        for j, valor in enumerate(row, start=1):
                            try:
                                valor_limpo = clean_cell_value(valor)
                                sheet_destino_phc.cell(row=i, column=j, value=valor_limpo)
                            except Exception as e:
                                print(f"Erro ao processar valor na linha {i}, coluna {j}")
                                print(f"Valor original: {valor}")
                                print(f"Valor após limpeza: {valor_limpo}")
                                raise e

                    wb_destino_phc.save(file_path_phc)
                    print(f"Dados PHC atualizados em: {file_path_phc}")

        # Criar arquivo de resumo
        create_summary_file(month_int, created_files, wb_bancos, dfs_phc)

    except Exception as e:
        print(f"Ocorreu um erro ao atualizar os ficheiros: {e}")
        print(f"Tipo do erro: {type(e)}")
        import traceback
        print(f"Detalhes do erro:\n{traceback.format_exc()}")

def create_summary_file(month_int, created_files, wb_bancos, dfs_phc):
    # Criar DataFrame para o resumo
    resumo_data = []

    for nome_banco, conta in mapeamento.items():
        # Buscar nome da coluna de valor para este banco
        coluna_valor = mapeamento_colunas_valor[nome_banco]

        # Calcular soma do banco
        if nome_banco in wb_bancos.sheetnames:
            sheet_banco = wb_bancos[nome_banco]
            dados_banco = pd.DataFrame(sheet_banco.values)
            # Encontrar índice da coluna pelo nome
            valor_col_idx = None
            for idx, cell in enumerate(dados_banco.iloc[0]):
                if cell == coluna_valor:
                    valor_col_idx = idx
                    break

            if valor_col_idx is not None:
                # Converter valores para numérico, ignorando cabeçalho
                valores_banco = pd.to_numeric(dados_banco.iloc[1:, valor_col_idx], errors='coerce')
                soma_banco = valores_banco.sum()
            else:
                soma_banco = 0
                print(f"Coluna '{coluna_valor}' não encontrada para {nome_banco}")
        else:
            soma_banco = 0
            print(f"Aba '{nome_banco}' não encontrada")

        # Calcular soma do PHC
        soma_phc = dfs_phc[conta]['Valor'].sum() if conta in dfs_phc else 0

        # Adicionar linha ao resumo
        resumo_data.append({
            'Conta': conta,
            'Nome': nome_banco,
            'Banco': soma_banco,
            'PHC': soma_phc,
            'Diferença': soma_banco - soma_phc
        })

    # Criar DataFrame de resumo
    df_resumo = pd.DataFrame(resumo_data)

    # Arredondar a coluna 'Diferença' para 2 casas decimais
    df_resumo['Diferença'] = df_resumo['Diferença'].round(2)

    # Opcional: também podemos arredondar as colunas 'Banco' e 'PHC'
    df_resumo['Banco'] = df_resumo['Banco'].round(2)
    df_resumo['PHC'] = df_resumo['PHC'].round(2)

    # Salvar em Excel
    month_name = meses_pt[month_int]
    output_file = f"Resumo Conciliação - {month_int} - {month_name}.xlsx"

    # Criar o relatório de documentos inválidos e salvar
    try:
        df_invalid_docs = create_invalid_docs_report(dfs_phc, 2024, month_int)

        # Salvar ambos os DataFrames no mesmo arquivo Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
            # Converter todas as colunas para string e limpar valores
            for col in df_invalid_docs.columns:
                df_invalid_docs[col] = df_invalid_docs[col].astype(str).apply(clean_cell_value)
            df_invalid_docs.to_excel(writer, sheet_name='Docs Inválidos', index=False)

    except Exception as e:
        print(f"Erro ao criar arquivo de resumo: {e}")
        print(f"Tipo do erro: {type(e)}")
        import traceback
        print(f"Detalhes do erro:\n{traceback.format_exc()}")

# Uso do script
if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python script.py <número_do_mês>")
    else:
        try:
            month_int = int(sys.argv[1])
            create_and_update_files(month_int)
        except ValueError:
            print("Por favor, insira um número válido para o mês.")